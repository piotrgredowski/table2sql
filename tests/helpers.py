import os
import tempfile
from typing import Any, Iterable, Optional, Tuple, cast

import openpyxl
from openpyxl.worksheet.worksheet import Worksheet


def save_to_csv(csv_str: str):
    with tempfile.NamedTemporaryFile(delete=False) as f:
        f.write(bytes(csv_str, encoding="utf-8"))
    new_filename = f.name + ".csv"
    os.rename(f.name, new_filename)

    return new_filename


def save_to_excel(
    rows: Iterable[Tuple[Any]],
    file_extension: str = "xlsx",
    sheet_name: Optional[str] = None,
    additional_sheets: Optional[Iterable[str]] = None,
):
    workbook = openpyxl.Workbook()

    if sheet_name:
        workbook.create_sheet(sheet_name)
        sheet = cast(Worksheet, workbook[sheet_name])
    else:
        sheet = workbook.active

    if additional_sheets:
        for additional_sheet in additional_sheets:
            workbook.create_sheet(additional_sheet)

    sheet.title = sheet_name or "some sheet"

    for r_idx, row in enumerate(rows, start=1):
        for c_idx, value in enumerate(row, start=1):
            _ = sheet.cell(column=c_idx, row=r_idx, value=value)

    with tempfile.NamedTemporaryFile(delete=False) as f:
        new_filename = f.name + "." + file_extension
        os.rename(f.name, new_filename)
        workbook.save(filename=new_filename)
    return new_filename
